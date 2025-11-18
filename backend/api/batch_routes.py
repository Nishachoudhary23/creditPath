from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from fastapi.responses import StreamingResponse
from backend.model.model_loader import load_model
from backend.utils.recommendation import get_recommendation
from backend.auth.routes import get_current_user
import pandas as pd
import io
import logging
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import PatternFill, Font

logger = logging.getLogger('CreditPathAI')
router = APIRouter()

@router.post("/predict_batch_file")
async def predict_batch_file(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
        
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        required_columns = ['loan_amnt', 'annual_inc', 'dti', 'open_acc', 'credit_age', 'revol_util']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing_columns)}"
            )
        
        model_data = load_model()
        model = model_data['model']
        scaler = model_data['scaler']
        
        results = []
        for idx, row in df.iterrows():
            financial_features = {
                'loan_amnt': row['loan_amnt'],
                'annual_inc': row['annual_inc'],
                'dti': row['dti'],
                'open_acc': row['open_acc'],
                'credit_age': row['credit_age'],
                'revol_util': row['revol_util']
            }
            
            df_row = pd.DataFrame([financial_features])
            df_scaled = scaler.transform(df_row)
            probability = float(model.predict_proba(df_scaled)[0][1])
            recommendation = get_recommendation(probability)
            
            results.append({
                'row_number': idx + 1,
                'probability': round(probability, 4),
                'risk_level': recommendation['risk_band'],
                'recommendation': recommendation['action'],
                **financial_features
            })
        
        logger.info(f"Batch prediction completed for {len(results)} rows")
        return {"predictions": results, "total": len(results)}
        
    except Exception as e:
        logger.error(f"Batch file prediction error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@router.post("/download_batch_results")
async def download_batch_results(predictions: dict, current_user: dict = Depends(get_current_user)):
    try:
        df = pd.DataFrame(predictions['predictions'])
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Predictions')
            workbook = writer.book
            worksheet = writer.sheets['Predictions']
            
            risk_col_index = df.columns.get_loc('risk_level') + 1 if 'risk_level' in df.columns else None
            if risk_col_index:
                low_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                medium_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                high_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                bold_font = Font(bold=True)

                for row_idx, value in enumerate(df['risk_level'], start=2):
                    cell = worksheet.cell(row=row_idx, column=risk_col_index)
                    risk_value = (value or '').strip().lower()
                    if risk_value == 'low':
                        cell.fill = low_fill
                    elif risk_value == 'medium':
                        cell.fill = medium_fill
                    elif risk_value == 'high':
                        cell.fill = high_fill
                    cell.font = bold_font

        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=batch_predictions.xlsx"}
        )
    except Exception as e:
        logger.error(f"Error creating download file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error creating download: {str(e)}")
